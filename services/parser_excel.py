from openpyxl import load_workbook

async def parse_excel(message):

    document = message.document

    file = await message.bot.get_file(document.file_id)

    await message.bot.download_file(file.file_path, "Project/uploads/excel/temp.xlsx")

    wb = load_workbook("Project/uploads/excel/temp.xlsx")
    sheet = wb.active

    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    await message.answer(
        f"Нашёл строк:\n{len(data)}"
    )
